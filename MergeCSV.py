#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
MergeCSV.py – Variante optimisée en gardant openpyxl
Objectif : réduire le temps total (lecture CSV + dédup + mise en forme) sans changer le rendu.

Optimisations clés :
- Lecture CSV plus rapide (engine "c" + détection simple du séparateur) avec fallback robuste.
- Déduplication via hash md5 de texte normalisé (au lieu d'une concaténation géante).
- Calcul des fusions consécutives (Question/Feedback) côté pandas, puis application en une passe sur la feuille.
- Évitations de certains recalculs.

NB : Les bordures/alignements restent appliqués cellule par cellule (limitation d'openpyxl),
mais l'essentiel du gain vient de la lecture, de la dédup et de la fusion en une passe.
"""

import os
import re
import csv
import hashlib
import unicodedata
from typing import List, Optional, Tuple

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side
from openpyxl.utils import get_column_letter

# =========================
# Chemins / paramètres
# =========================
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
INPUT_DIR = os.environ.get("INPUT_FOLDER", SCRIPT_DIR)
OUTPUT_DIR = os.environ.get("OUTPUT_FOLDER", os.path.join(SCRIPT_DIR, "merged_files"))
os.makedirs(OUTPUT_DIR, exist_ok=True)

FINAL_YEAR = os.environ.get("FINAL_YEAR") or "2025"

# Colonnes retirées de la feuille QUIZ
DROP_COLUMNS = {"Numéro", "Nom", "Importante", "source_fichier"}

# Largeurs de colonnes EXACTES demandées
FIXED_COLUMN_WIDTHS = {
    "Question": 50,
    "Réponse": 30,
    "Valide": 10,
    "Feedback": 140,
    "NbCar Feedback": 16,
}

# Colonnes à fusionner visuellement si identiques consécutives
MERGE_VISUAL_COLS = ["Question", "Feedback"]

# Hauteur des lignes en points
ROW_HEIGHT_POINTS = 150.0


# =========================
# Utilitaires
# =========================

def _normalize_text(s: Optional[str]) -> str:
    if s is None:
        return ""
    s = unicodedata.normalize("NFKC", str(s))
    s = (
        s.replace("’", "'")
        .replace("‘", "'")
        .replace("“", '"')
        .replace("”", '"')
        .replace("–", "-")
        .replace("—", "-")
        .replace("\u00A0", " ")
    )
    s = re.sub(r"\s+", " ", s).strip()
    return s


def _detect_sep_from_sample(sample: bytes) -> str:
    """Détecte un séparateur probable (, ; \t) sur la première ligne du fichier."""
    try:
        first_line = sample.decode("utf-8", errors="ignore").splitlines()[0]
    except Exception:
        return ","
    best_sep, best_cols = ",", 1
    for sep in (",", ";", "\t"):
        cols = len(first_line.split(sep))
        if cols > best_cols:
            best_sep, best_cols = sep, cols
    return best_sep


def read_csv_fast(path: str) -> pd.DataFrame:
    """Lecture CSV rapide : engine='c' + encodages usuels, fallback python/autodetect en dernier."""
    encodings = ("utf-8-sig", "utf-8", "latin-1", "cp1252")
    with open(path, "rb") as f:
        sample = f.read(4096)
    sep = _detect_sep_from_sample(sample)

    last_err = None
    for enc in encodings:
        try:
            return pd.read_csv(
                path,
                sep=sep,
                engine="c",
                encoding=enc,
                dtype=str,
                na_filter=False,
                quoting=csv.QUOTE_MINIMAL,
                low_memory=False,
            )
        except Exception as e:
            last_err = e
            continue

    # Fallback tolérant
    for enc in encodings:
        try:
            return pd.read_csv(path, sep=None, engine="python", encoding=enc, dtype=str)
        except Exception as e:
            last_err = e
    raise ValueError(f'Lecture CSV "{os.path.basename(path)}" impossible : {last_err}')


def sanitize_filename(name: str) -> str:
    return re.sub(r'[\\/:*?"<>|]', " ", name).strip()


def build_final_name_from_content(df_first: pd.DataFrame) -> str:
    numero = None
    if "Numéro" in df_first.columns:
        s = (
            pd.Series(df_first["Numéro"], copy=False)
            .astype(str)
            .str.strip()
            .replace("", pd.NA)
            .dropna()
        )
        if not s.empty:
            numero = s.iloc[0]
    if not numero:
        numero = "0000"
    m = re.match(r"^(\d{4})", str(numero))
    numero_clean = m.group(1) if m else str(numero).strip()

    titre = None
    if "Nom" in df_first.columns:
        s = (
            pd.Series(df_first["Nom"], copy=False)
            .astype(str)
            .str.strip()
            .replace("", pd.NA)
            .dropna()
        )
        if not s.empty:
            titre = s.iloc[0]
    if not titre:
        titre = "quiz"
    titre = re.sub(r"\s+", " ", str(titre)).strip()

    final_name = f"{numero_clean}_{titre}_biblio_{FINAL_YEAR}.xlsx"
    return sanitize_filename(final_name)


def normalize_df_for_key(df: pd.DataFrame) -> pd.DataFrame:
    df2 = df.copy()
    for c in df2.columns:
        df2[c] = df2[c].astype(str).map(_normalize_text).str.lower()
    return df2


def make_hash_key(df_norm: pd.DataFrame) -> pd.Series:
    def row_hash(values: Tuple[str, ...]) -> str:
        h = hashlib.md5()
        for v in values:
            h.update(v.encode("utf-8", errors="ignore"))
            h.update(b"||")
        return h.hexdigest()

    return pd.Series(
        (row_hash(tuple(values)) for values in df_norm.itertuples(index=False, name=None)),
        index=df_norm.index,
        dtype="object",
    )


def compute_merge_runs(df: pd.DataFrame, col_name: str) -> List[Tuple[int, int]]:
    """Retourne des (row_start, row_end) 1-based (hors entête) pour les valeurs consécutives identiques non vides."""
    if col_name not in df.columns or df.empty:
        return []
    vals = df[col_name].astype(str).map(_normalize_text).str.lower().tolist()
    runs = []
    start = 0
    for i in range(1, len(vals) + 1):
        prev = vals[i - 1] if i - 1 < len(vals) else None
        cur = vals[i] if i < len(vals) else None
        if i == len(vals) or cur != prev or prev == "":
            length = i - start
            if length > 1 and prev != "":
                # +2 : passage en index feuille (1 = entête, 2 = 1ère ligne de données)
                runs.append((start + 2, i + 1))
            start = i
    return runs


# =========================
# Écriture / Mise en forme (openpyxl)
# =========================

def write_quiz_sheet(wb: Workbook, df: pd.DataFrame) -> None:
    ws = wb.active
    ws.title = "QUIZ"

    # 1) Entête + données
    headers = list(df.columns)
    ws.append(headers)
    for row in df.itertuples(index=False, name=None):
        ws.append(row)

    header_map = {name: i + 1 for i, name in enumerate(headers)}

    # 2) Gel + filtre
    ws.freeze_panes = "A2"
    last_col_letter = get_column_letter(ws.max_column)
    ws.auto_filter.ref = f"A1:{last_col_letter}1"

    # 3) Hauteurs & alignements
    align_wrap_left = Alignment(horizontal='left', vertical='center', wrap_text=True)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        ws.row_dimensions[row[0].row].height = ROW_HEIGHT_POINTS
        for cell in row:
            if cell.value is not None:
                cell.alignment = align_wrap_left

    # 4) Largeurs fixes
    for col_name, width in FIXED_COLUMN_WIDTHS.items():
        idx = header_map.get(col_name)
        if idx:
            ws.column_dimensions[get_column_letter(idx)].width = width

    # 5) Bordures fines
    thin = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thin

    # 6) Fusions visuelles – pré-calculées
    for col in MERGE_VISUAL_COLS:
        idx = header_map.get(col)
        if not idx:
            continue
        for r1, r2 in compute_merge_runs(df, col):
            ws.merge_cells(start_row=r1, end_row=r2, start_column=idx, end_column=idx)
            ws.cell(row=r1, column=idx).alignment = align_wrap_left

    # 7) Formule LEN(Feedback) -> "NbCar Feedback"
    c_feedback = header_map.get("Feedback")
    c_len = header_map.get("NbCar Feedback")
    if c_feedback and c_len:
        colF = get_column_letter(c_feedback)
        for r in range(2, ws.max_row + 1):
            ws.cell(row=r, column=c_len).value = f"=LEN({colF}{r})"


# =========================
# Pipeline principal
# =========================

def main():
    csv_files: List[str] = sorted([f for f in os.listdir(INPUT_DIR) if f.lower().endswith(".csv")])
    if not csv_files:
        print(f"Aucun CSV trouvé dans {INPUT_DIR}")
        return

    # Nom final d'après contenu du 1er CSV
    df_first = read_csv_fast(os.path.join(INPUT_DIR, csv_files[0]))
    final_name = build_final_name_from_content(df_first)
    final_path = os.path.join(OUTPUT_DIR, final_name)

    # Nettoyer anciens .xlsx
    for name in os.listdir(OUTPUT_DIR):
        p = os.path.join(OUTPUT_DIR, name)
        if os.path.isfile(p) and name.lower().endswith(".xlsx"):
            try:
                os.remove(p)
            except Exception:
                pass

    # Fusion
    frames: List[pd.DataFrame] = [df_first]
    for name in csv_files[1:]:
        frames.append(read_csv_fast(os.path.join(INPUT_DIR, name)))
    fusion = pd.concat(frames, axis=0, ignore_index=True, sort=False)

    # Retirer colonnes inutiles
    fusion = fusion[[c for c in fusion.columns if c not in DROP_COLUMNS]]

    # Dédup rapide
    norm = normalize_df_for_key(fusion)
    key = make_hash_key(norm)
    fusion = fusion.loc[~key.duplicated(keep="first")].reset_index(drop=True)

    # Ajout NbCar Feedback si besoin
    if "Feedback" in fusion.columns and "NbCar Feedback" not in fusion.columns:
        fusion["NbCar Feedback"] = 0

    # Ordre colonnes
    preferred = ["Question", "Type de question", "Réponse", "Valide", "Feedback", "NbCar Feedback"]
    fusion = fusion[[c for c in preferred if c in fusion.columns] + [c for c in fusion.columns if c not in preferred]]

    # Écriture finale : QUIZ + BIBLIOGRAPHIE
    wb = Workbook()
    write_quiz_sheet(wb, fusion)
    wb.create_sheet(title="BIBLIOGRAPHIE")
    wb.save(final_path)
    wb.close()

    os.environ["FINAL_XLSX_NAME"] = final_name
    print("✅ Fichier final généré :", final_path)


if __name__ == "__main__":
    main()
