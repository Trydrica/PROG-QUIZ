#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import os
from datetime import datetime
from typing import Dict

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter

# -------------------------------------------------------------------
# RÉPERTOIRE DE TRAVAIL
# -------------------------------------------------------------------
# Si Main.py a défini OUTPUT_FOLDER, on l'utilise ; sinon fallback "merged_files".
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
DIRECTORY = os.environ.get("OUTPUT_FOLDER", os.path.join(SCRIPT_DIR, "merged_files"))

if not os.path.exists(DIRECTORY):
    raise FileNotFoundError(f"Le dossier '{DIRECTORY}' n'existe pas.")

# -------------------------------------------------------------------
# PARAMÈTRES DE PRÉSENTATION (adapter librement à tes besoins)
# -------------------------------------------------------------------
COLUMN_WIDTHS: Dict[str, float] = {
    "Numéro": 10,
    "Nom": 30,
    "Question": 140,
    "Type de question": 24,
    "Réponse": 70,
    "Valide": 10,
    "Importante": 14,
    "Feedback": 140,
    "source_fichier": 28,
}

WRAP_COLS = {"Question", "Réponse", "Feedback"}

BORDER_THIN = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

HEADER_FONT = Font(bold=True)

# -------------------------------------------------------------------
# OUTILS
# -------------------------------------------------------------------
def format_sheet(ws, title: str = None) -> None:
    """Applique le format commun : header figé, auto-filtre, largeurs, wraps, bordures légères."""
    # Figer l’entête
    ws.freeze_panes = "A2"

    # Auto-filtre
    last_col_letter = get_column_letter(ws.max_column)
    ws.auto_filter.ref = f"A1:{last_col_letter}1"

    # Titre facultatif en A1 (si tu veux remplacer l'en-tête par un titre)
    if title:
        ws["A1"].value = title
        ws["A1"].font = HEADER_FONT

    # Récupère l'entête pour faire correspondre les largeurs
    headers = [c.value if c.value is not None else "" for c in ws[1]]
    header_to_idx = {str(name): i + 1 for i, name in enumerate(headers)}

    # Largeurs
    for col_name, width in COLUMN_WIDTHS.items():
        idx = header_to_idx.get(col_name)
        if idx:
            ws.column_dimensions[get_column_letter(idx)].width = width

    # Wrap text + align top sur colonnes longues
    for col_name in WRAP_COLS:
        idx = header_to_idx.get(col_name)
        if idx:
            for row in ws.iter_rows(min_row=2, min_col=idx, max_col=idx):
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Bordures fines sur toutes les cellules remplies (facultatif)
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = BORDER_THIN

def touch_updated_copy(xlsx_path: str) -> str:
    """
    Sauvegarde une copie avec suffixe _updated (ou remplace le fichier si tu préfères).
    Retourne le chemin du nouveau fichier.
    """
    base, ext = os.path.splitext(xlsx_path)
    new_path = base + "_updated" + ext
    return new_path

# -------------------------------------------------------------------
# BOUCLE PRINCIPALE
# -------------------------------------------------------------------
def main() -> None:
    xlsx_files = [f for f in os.listdir(DIRECTORY) if f.lower().endswith(".xlsx")]
    if not xlsx_files:
        print(f"Aucun .xlsx à transformer dans {DIRECTORY}")
        return

    for name in sorted(xlsx_files):
        in_path = os.path.join(DIRECTORY, name)
        out_path = touch_updated_copy(in_path)

        # Charger le classeur
        wb = load_workbook(in_path)
        # Si plusieurs feuilles, on les traite toutes
        for ws in wb.worksheets:
            # Option : titre contextuel (ex. nom du fichier + date)
            title = None  # ou f"{name} – {datetime.now():%Y-%m-%d}"
            format_sheet(ws, title=title)

        # Sauvegarde
        wb.save(out_path)
        wb.close()
        print(f"✅ Fichier formaté : {out_path}")

    print("Tous les fichiers ont été traités.")

if __name__ == "__main__":
    main()