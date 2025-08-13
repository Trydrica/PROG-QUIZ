#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Main.py — Pipeline CSV -> Excel pour le site

Fonctions :
- Convertit chaque CSV de input_folder en .xlsx dans output_folder
- Crée une fusion globale (fusion_globale.xlsx)
- Crée des fusions par groupe 10/20/30... si le nom du fichier contient un code à 4 chiffres (ex. 1001 -> groupe 10)

Utilisation :
    python Main.py <input_folder> <output_folder>

Peut aussi être importé :
    from Main import process_folder
    process_folder("/path/in", "/path/out")
"""

import os
import re
import sys
import csv
import shutil
from io import BytesIO, StringIO
from typing import List, Optional, Tuple

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter


# ----------------------------
# Utils
# ----------------------------

def ensure_empty_dir(path: str) -> None:
    """Crée le dossier s'il n'existe pas, et le vide sinon."""
    os.makedirs(path, exist_ok=True)
    # On NE supprime que des fichiers xlsx/csv connus pour éviter un wipe trop agressif
    for name in os.listdir(path):
        full = os.path.join(path, name)
        if os.path.isfile(full) and (name.lower().endswith((".xlsx", ".csv"))):
            try:
                os.remove(full)
            except Exception:
                pass


def read_csv_robust(path: str) -> pd.DataFrame:
    """
    Lecture robuste d'un CSV :
    - Détection automatique du séparateur (',' ';' '\t' …) via engine='python' + sep=None
    - Tentatives d'encodage : utf-8-sig, utf-8, latin-1, cp1252
    - dtype=str pour préserver les zéros initiaux et les identifiants
    """
    last_err: Optional[Exception] = None
    for enc in ("utf-8-sig", "utf-8", "latin-1", "cp1252"):
        try:
            df = pd.read_csv(path, sep=None, engine="python", encoding=enc, dtype=str)
            return df
        except Exception as e:
            last_err = e
    raise ValueError(f'Lecture CSV "{os.path.basename(path)}" impossible : {last_err}')


def extract_group_from_filename(filename: str) -> Optional[int]:
    """
    Extrait un code groupe 10/20/30… si le nom contient 4 chiffres d'affilée.
    Ex.: 'quiz-1001.csv' -> 10 ; 'truc_2033.csv' -> 20 ; sinon None
    """
    m = re.search(r"(\d{4})", filename)
    return int(m.group(1)[:2]) if m else None


def write_xlsx_with_format(df: pd.DataFrame, out_path: str, sheet_name: str = "Données") -> None:
    """
    Écrit un DataFrame en .xlsx + quelques mises en forme utiles :
    - Fige l'entête (freeze A2)
    - Active l'auto-filtre sur l'entête
    - Largeurs de colonnes pré-définies si les titres correspondent
    - Wrap text pour les colonnes longues (Question/Réponse/Feedback)
    """
    # Écriture initiale
    with pd.ExcelWriter(out_path, engine="openpyxl") as w:
        df.to_excel(w, index=False, sheet_name=sheet_name)

    # Mises en forme
    wb = load_workbook(out_path)
    ws = wb[sheet_name]

    # En-tête figée et filtre
    ws.freeze_panes = "A2"
    last_col_letter = get_column_letter(ws.max_column)
    ws.auto_filter.ref = f"A1:{last_col_letter}1"

    # Largeurs de colonnes possibles (ajuste si tes entêtes diffèrent)
    # Adapter les clés à tes entêtes exacte (casse/accents compris)
    column_widths = {
        "Numéro": 10,
        "Nom": 30,
        "Question": 140,
        "Type de question": 24,
        "Réponse": 70,
        "Valide": 10,
        "Importante": 14,
        "Feedback": 140,
        # ajoute ici si besoin...
    }

    # Applique les largeurs si le titre existe
    headers = [cell.value if cell.value is not None else "" for cell in ws[1]]
    header_to_idx = {str(name): i + 1 for i, name in enumerate(headers)}

    for col_name, width in column_widths.items():
        idx = header_to_idx.get(col_name)
        if idx:
            ws.column_dimensions[get_column_letter(idx)].width = width

    # Wrap text sur colonnes longues
    wrap_cols = {"Question", "Réponse", "Feedback"}
    for col_name in wrap_cols:
        idx = header_to_idx.get(col_name)
        if idx:
            for row in ws.iter_rows(min_row=2, min_col=idx, max_col=idx):
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")

    wb.save(out_path)
    wb.close()


# ----------------------------
# Coeur du pipeline
# ----------------------------

def process_folder(input_folder: str, output_folder: str) -> Tuple[List[str], List[str]]:
    """
    Traite tous les CSV de input_folder et écrit les XLSX dans output_folder.
    Retourne (liste_xlsx_individuels, liste_xlsx_fusion).
    """
    if not os.path.isdir(input_folder):
        raise NotADirectoryError(f"input_folder invalide : {input_folder}")

    ensure_empty_dir(output_folder)

    csv_files = [f for f in os.listdir(input_folder) if f.lower().endswith(".csv")]
    if not csv_files:
        raise FileNotFoundError(f"Aucun .csv trouvé dans : {input_folder}")

    global_dfs: List[pd.DataFrame] = []
    per_group: dict[int, List[pd.DataFrame]] = {}
    written_individuals: List[str] = []
    written_fusions: List[str] = []

    for name in sorted(csv_files):
        src_path = os.path.join(input_folder, name)
        df = read_csv_robust(src_path)

        # Conserver la source
        df.insert(0, "source_fichier", name)

        # 1) Excel individuel
        base = os.path.splitext(name)[0]
        out_indiv = os.path.join(output_folder, f"{base}.xlsx")
        write_xlsx_with_format(df, out_indiv, sheet_name="Données")
        written_individuals.append(out_indiv)
        print(f"✅ Fichier Excel créé : {out_indiv}")

        # 2) alimenter la fusion globale
        global_dfs.append(df)

        # 3) alimenter la fusion par groupe (si applicable)
        grp = extract_group_from_filename(name)
        if grp is not None:
            per_group.setdefault(grp, []).append(df)

    # 4) Fusion globale
    if global_dfs:
        fusion_all = pd.concat(global_dfs, axis=0, ignore_index=True, sort=False)
        out_fusion = os.path.join(output_folder, "fusion_globale.xlsx")
        write_xlsx_with_format(fusion_all, out_fusion, sheet_name="Fusion")
        written_fusions.append(out_fusion)
        print(f"✅ Fusion globale : {out_fusion}")

    # 5) Fusions par groupe
    for grp, dfs in sorted(per_group.items()):
        fusion_g = pd.concat(dfs, axis=0, ignore_index=True, sort=False)
        out_grp = os.path.join(output_folder, f"Group_{grp}.xlsx")
        write_xlsx_with_format(fusion_g, out_grp, sheet_name=f"Groupe{grp}")
        written_fusions.append(out_grp)
        print(f"✅ Fusion groupe {grp:02d} : {out_grp}")

    return written_individuals, written_fusions


# ----------------------------
# Entrée CLI
# ----------------------------

def main():
    if len(sys.argv) != 3:
        print("Utilisation : python Main.py <input_folder> <output_folder>")
        sys.exit(1)

    input_folder = os.path.abspath(sys.argv[1])
    output_folder = os.path.abspath(sys.argv[2])

    print(f"➡️  Input : {input_folder}")
    print(f"➡️  Output: {output_folder}")

    try:
        indiv, fusions = process_folder(input_folder, output_folder)
        print("\n--- RÉSUMÉ ---")
        for p in indiv:
            print(f"[INDIV]  {p}")
        for p in fusions:
            print(f"[FUSION] {p}")
        print("\nTerminé ✅")
    except Exception as e:
        print("❌ Erreur :", e)
        sys.exit(2)


if __name__ == "__main__":
    main()